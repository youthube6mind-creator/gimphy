from dataclasses import asdict, dataclass
from enum import Enum, auto
from ftfcu_appworx import Apwx, JobTime
from oracledb import Connection as DbConnection
from pathlib import Path
from typing import Any, Optional, List, Dict, Tuple

import csv
import datetime
import openpyxl
import yaml
import os
import sys

__version__ = 1.07

# TODO: Update description on github
# TODO: Update permissions on github
# TODO: Update readme
# TODO: Update config
# TODO: Update requirements.txt
# TODO: Update script version
# TODO: CICD setup
# TODO: Create automated unit tests
# TODO: Add comments
# TODO: Add error handling


class AppWorxEnum(Enum):
    """Define AppWorx arguments here to avoid hard-coded strings."""

    HOST = auto()
    SID = auto()
    P2P_SERVER = auto()
    P2P_SCHEMA = auto()
    CONFIG_FILE = auto()
    INPUT_FILE = auto()
    OUTPUT_FILE_PATH = auto()
    RPT_ONLY = auto()
    RECONCILE_DATE = auto()

    def __str__(self):
        return self.name


@dataclass
class ScriptData:
    """Class that holds all the structures and data needed by the script."""
    
    apwx: Apwx
    p2p_dbh: DbConnection
    dna_dbh: DbConnection
    config: Any


def run(apwx: Apwx):
    """Main processing function for the script."""
    
    print(f"P2P Manual Reconcile job started at {datetime.datetime.now()}")
    
    script_data = initialize(apwx)
    
    input_file_path = apwx.args.INPUT_FILE
    
    print("Processing Reconciliation File")
    trans_to_reconcile = parse_recon_file(input_file_path, script_data.config)
    
    # Process reconcile date format
    reconcile_date = getattr(apwx.args, 'RECONCILE_DATE', None)
    if reconcile_date:
        reconcile_date = reconcile_date.replace('-', '/')
    
    print("Opening output file")
    output_file_path = create_output_file_path(apwx.args.OUTPUT_FILE_PATH)
    
    with open(output_file_path, 'w') as fh_out:
        write_output_header(fh_out)
        
        print("Updating Reconcile Dates")
        update_reconcile_date(
            script_data, 
            fh_out, 
            trans_to_reconcile, 
            reconcile_date, 
            apwx.args.RPT_ONLY
        )
    
    print("Disconnecting databases")
    script_data.p2p_dbh.close()
    script_data.dna_dbh.close()
    
    print(f"P2P Manual Reconcile job finished at {datetime.datetime.now()}")
    
    return True


def get_apwx() -> Apwx:
    """Creates the appworx object."""
    return Apwx(["OSIUPDATE", "OSIUPDATE_PW"])


def parse_args(apwx: Apwx) -> Apwx:
    """Validates the parameters provided to the script."""
    parser = apwx.parser
    parser.add_arg(str(AppWorxEnum.HOST), type=str, required=True)
    parser.add_arg(str(AppWorxEnum.SID), type=str, required=True)
    parser.add_arg(str(AppWorxEnum.P2P_SERVER), type=str, required=True)
    parser.add_arg(str(AppWorxEnum.P2P_SCHEMA), type=str, required=True)
    parser.add_arg(
        str(AppWorxEnum.CONFIG_FILE), type=r"(\.yml|\.yaml)$", required=True
    )
    parser.add_arg(
        str(AppWorxEnum.INPUT_FILE), type=r"\.xlsx$", required=True
    )
    parser.add_arg(
        str(AppWorxEnum.OUTPUT_FILE_PATH), type=parser.dir_validator, required=True
    )
    parser.add_arg(
        str(AppWorxEnum.RPT_ONLY), choices=["Y", "N"], default="N", required=True
    )
    parser.add_arg(
        str(AppWorxEnum.RECONCILE_DATE), type=r"^\d{2}-\d{2}-\d{4}$", required=False
    )
    apwx.parse_args()
    return apwx


def p2p_db_connect(apwx) -> DbConnection:
    """Creates the P2P database connection object."""
    # This would use P2P specific connection details
    # For now, using the standard DNA connection as placeholder
    return apwx.db_connect(autocommit=False)


def dna_db_connect(apwx) -> DbConnection:
    """Creates the DNA database connection object."""
    return apwx.db_connect(autocommit=False)


def get_config(apwx: Apwx) -> dict:
    """Loads config YAML into a dictionary."""
    with open(apwx.args.CONFIG_FILE, "r") as f:
        return yaml.safe_load(f)


def initialize(apwx) -> ScriptData:
    """Initialize objects required by the script to call external systems."""
    p2p_dbh = p2p_db_connect(apwx)
    dna_dbh = dna_db_connect(apwx)
    config = get_config(apwx)
    
    # Set autocommit to False
    p2p_dbh.autocommit = config['database_config']['p2p_autocommit']
    dna_dbh.autocommit = config['database_config']['dna_autocommit']
    
    return ScriptData(apwx=apwx, p2p_dbh=p2p_dbh, dna_dbh=dna_dbh, config=config)


def parse_recon_file(input_file_path: str, config: dict) -> List[Dict]:
    """Parse the Excel reconciliation file and return transaction data."""
    workbook = openpyxl.load_workbook(input_file_path)
    worksheet = workbook.active
    
    valid_col_headers = config['valid_column_headers']
    
    # Get headers from first row
    col_headers = []
    for col in range(1, len(valid_col_headers) + 1):
        cell_value = worksheet.cell(row=1, column=col).value
        col_headers.append(cell_value)
    
    # Validate headers
    for i, header in enumerate(col_headers):
        if i < len(valid_col_headers) and header != valid_col_headers[i]:
            raise ValueError(f"{header} is not a valid column name or is in the wrong position")
    
    # Parse data rows
    file_data = []
    max_row = worksheet.max_row
    
    for row_num in range(2, max_row + 1):  # Skip header row
        row_data = {}
        for col_num, header in enumerate(valid_col_headers, 1):
            cell_value = worksheet.cell(row=row_num, column=col_num).value
            row_data[header] = cell_value
        file_data.append(row_data)
    
    return file_data


def create_output_file_path(output_dir: str) -> str:
    """Create the output file path with timestamp."""
    if not output_dir.endswith('\\') and not output_dir.endswith('/'):
        output_dir += os.sep
    
    today = datetime.date.today().strftime('%m-%d-%Y')
    filename = f"P2P_RECON_MANUAL_UPDATE_{today}.txt"
    
    return os.path.join(output_dir, filename)


def write_output_header(fh_out):
    """Write the header information to the output file."""
    fh_out.write("P2P RECON MANUAL UPDATE\n")
    fh_out.write(f"RUN DATE: {datetime.datetime.now()}\n")
    fh_out.write("-" * 150 + "\n")
    fh_out.write("\n")


def execute_sql(conn: DbConnection, sql_statement: str, params: List = None) -> Tuple[int, str]:
    """Execute SQL statement and return number of affected rows and error message."""
    try:
        with conn.cursor() as cursor:
            if params:
                cursor.execute(sql_statement, params)
            else:
                cursor.execute(sql_statement)
            
            rows_affected = cursor.rowcount
            return rows_affected, None
    except Exception as e:
        return 0, str(e)


def update_p2p_recon_date(dbh: DbConnection, p2p_type: str, record_id: int, 
                         reconcile_date: str, config: dict) -> Tuple[int, str]:
    """Update reconciliation date in P2P tables."""
    if p2p_type == 'FTF':
        sql = config['sql_queries']['update_payment']
    else:
        sql = config['sql_queries']['update_detail_record']
    
    if reconcile_date:
        params = [reconcile_date, record_id]
    else:
        # Use current date
        current_date = datetime.datetime.now().strftime('%m/%d/%Y')
        params = [current_date, record_id]
    
    return execute_sql(dbh, sql, params)


def update_rtxn_recon_date(dbh: DbConnection, acctnbr: int, rtxnnbr: int, 
                          reconcile_date: str, config: dict) -> Tuple[int, str]:
    """Update reconciliation date in RTXN entity attributes."""
    sql = config['sql_queries']['insert_rtxn_recon_date']
    
    if reconcile_date:
        date_value = reconcile_date
    else:
        date_value = datetime.datetime.now()
    
    params = [acctnbr, rtxnnbr, date_value, acctnbr, rtxnnbr]
    
    return execute_sql(dbh, sql, params)


def update_card_recon_date(dbh: DbConnection, network_id: str, network_tran_id: str, 
                          tran_code: str, reconcile_date: str, config: dict) -> Tuple[int, str]:
    """Update reconciliation date in card network tables."""
    if network_id == 'MC':
        sql = config['sql_queries']['update_mc_recon']
        if reconcile_date:
            params = [reconcile_date, network_tran_id]
        else:
            params = [datetime.date.today(), network_tran_id]
    else:  # VISA
        sql = config['sql_queries']['update_visa_recon']
        if reconcile_date:
            params = [reconcile_date, network_tran_id, tran_code]
        else:
            params = [datetime.date.today(), network_tran_id, tran_code]
    
    return execute_sql(dbh, sql, params)


def update_reconcile_date(script_data: ScriptData, fh_out, trans_to_reconcile: List[Dict], 
                         reconcile_date: str, rpt_only: str):
    """Process each transaction and update reconciliation dates."""
    row_ct = 1  # Start count includes header row
    
    if not reconcile_date:
        reconcile_date = datetime.date.today().strftime('%m/%d/%Y')
    
    for tran in trans_to_reconcile:
        row_ct += 1
        
        # Write transaction details
        tran_line = "\n".join([f"{k}: {v if v else 'N/A'}" for k, v in sorted(tran.items())])
        
        fh_out.write(f"Row {row_ct}:\n")
        fh_out.write("-" * 75 + "\n")
        fh_out.write(tran_line + "\n")
        
        # Update Detail Record
        if tran.get('UPDATE_DETAIL_RECORD', '').upper() == 'Y':
            fh_out.write("DetailRecord Table Update Status:\n")
            
            detail_record_id = tran.get('DETAIL_RECORD_ID')
            if not detail_record_id or not str(detail_record_id).isdigit():
                fh_out.write("Reconcile Date Not Updated: DetailRecord Id is undefined or non numeric\n")
            else:
                updated, err_msg = update_p2p_recon_date(
                    script_data.p2p_dbh, 'ZELLE', int(detail_record_id), reconcile_date, script_data.config
                )
                
                if rpt_only == 'N':
                    script_data.p2p_dbh.commit()
                else:
                    script_data.p2p_dbh.rollback()
                
                if updated:
                    msg = ("Reconcile Date Updated: " + reconcile_date if updated > 0 
                          else "Reconcile Date Not Updated: Reconcile Date is already populated")
                    fh_out.write(msg + "\n")
                else:
                    fh_out.write(f"Reconcile Date Not Updated: {err_msg}\n")
        
        # Update Payment
        if tran.get('UPDATE_PAYMENT', '').upper() == 'Y':
            fh_out.write("Payment Table Update Status:\n")
            
            payment_id = tran.get('PAYMENT_ID')
            if not payment_id or not str(payment_id).isdigit():
                fh_out.write("Reconcile Date Not Updated: Payment Id is undefined or non numeric\n")
            else:
                updated, err_msg = update_p2p_recon_date(
                    script_data.p2p_dbh, 'FTF', int(payment_id), reconcile_date, script_data.config
                )
                
                if rpt_only == 'N':
                    script_data.p2p_dbh.commit()
                else:
                    script_data.p2p_dbh.rollback()
                
                if updated:
                    msg = ("Reconcile Date Updated: " + reconcile_date if updated > 0 
                          else "Reconcile Date Not Updated: Reconcile Date is already populated")
                    fh_out.write(msg + "\n")
                else:
                    fh_out.write(f"Reconcile Date Not Updated: {err_msg}\n")
        
        # Update RTXN
        if tran.get('UPDATE_RTXN', '').upper() == 'Y':
            fh_out.write("RTXN Entity Attribute Update Status:\n")
            
            acctnbr = tran.get('ACCTNBR')
            rtxnnbr = tran.get('RTXNNBR')
            
            if (not acctnbr or not str(acctnbr).isdigit() or 
                not rtxnnbr or not str(rtxnnbr).isdigit()):
                
                if not acctnbr or not str(acctnbr).isdigit():
                    fh_out.write("Reconcile Date Not Updated: ACCTNBR is undefined or non-numeric\n")
                if not rtxnnbr or not str(rtxnnbr).isdigit():
                    fh_out.write("Reconcile Date Not Updated: RTXNNBR is undefined or non-numeric\n")
            else:
                updated, err_msg = update_rtxn_recon_date(
                    script_data.dna_dbh, int(acctnbr), int(rtxnnbr), reconcile_date, script_data.config
                )
                
                if rpt_only == 'N':
                    script_data.dna_dbh.commit()
                else:
                    script_data.dna_dbh.rollback()
                
                if updated:
                    msg = ("Reconcile Date Updated: " + reconcile_date if updated > 0 
                          else "Reconcile Date Not Updated: Reconcile Date is already populated")
                    fh_out.write(msg + "\n")
                else:
                    fh_out.write(f"Reconcile Date Not Updated: {err_msg}\n")
        
        # Update MC
        if tran.get('UPDATE_MC', '').upper() == 'Y':
            fh_out.write("MC_ZELDLY Recon Table Update:\n")
            
            network_id = tran.get('NETWORK_ID')
            if not network_id:
                fh_out.write("Reconcile Date Not Updated: NETWORK_ID is undefined\n")
            else:
                updated, err_msg = update_card_recon_date(
                    script_data.dna_dbh, 'MC', network_id, None, reconcile_date, script_data.config
                )
                
                if rpt_only == 'N':
                    script_data.dna_dbh.commit()
                else:
                    script_data.dna_dbh.rollback()
                
                if updated:
                    msg = ("Reconcile Date Updated: " + reconcile_date if updated > 0 
                          else "Reconcile Date Not Updated: Reconcile Date is already populated")
                    fh_out.write(msg + "\n")
                else:
                    fh_out.write(f"Reconcile Date Not Updated: {err_msg}\n")
        
        # Update VISA
        if tran.get('UPDATE_VISA', '').upper() == 'Y':
            fh_out.write("VISA_RW3 Recon Table Update:\n")
            
            network_id = tran.get('NETWORK_ID')
            tran_type = tran.get('TRAN_TYPE')
            
            if (not network_id or not tran_type or 
                tran_type.upper() not in ['CREDIT', 'DEBIT']):
                
                if not network_id:
                    fh_out.write("Reconcile Date Not Updated: NETWORK_ID is undefined\n")
                if not tran_type or tran_type.upper() not in ['CREDIT', 'DEBIT']:
                    fh_out.write("Reconcile Date Not Updated: TRAN_CODE is undefined or not CREDIT or DEBIT\n")
            else:
                updated, err_msg = update_card_recon_date(
                    script_data.dna_dbh, 'VISA', network_id, tran_type, reconcile_date, script_data.config
                )
                
                if rpt_only == 'N':
                    script_data.dna_dbh.commit()
                else:
                    script_data.dna_dbh.rollback()
                
                if updated:
                    msg = ("Reconcile Date Updated: " + reconcile_date if updated > 0 
                          else "Reconcile Date Not Updated: Reconcile Date is already populated")
                    fh_out.write(msg + "\n")
                else:
                    fh_out.write(f"Reconcile Date Not Updated: {err_msg}\n")
        
        fh_out.write("\n")
    
    return True


if __name__ == "__main__":
    JobTime().print_start()
    run(parse_args(get_apwx()))
    JobTime().print_end()